# excel.py

import os
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.config import BASE_DIR


from .models import Country

class Excel:
    def __init__(self):
        self.name = "Keywords_Report.xlsx"
        self.output_folder_name = "output"
        self.output_path = BASE_DIR / self.output_folder_name
        self.workbook: Workbook = Workbook()
        self.countries: list[Country] = []
        self.create_output_path_if_not()
        
    def create_output_path_if_not(self):
        if not os.path.exists(self.output_path):
            os.makedirs(self.output_path)
            
    def create_sheets(self, headers: dict, keywords_manager):
        for country in headers.keys():
            if country == "Sheet1":
                continue
            sheet: Worksheet = self.workbook.create_sheet(title=country)
            current_column_index: int = 0
            print(f"Country: {country}")
            keywords = keywords_manager.database.get_keywords(country)
            # print(f"[Create Sheets] Getting Keywords For Country: {country}, Found {len(keywords)} Keywords")
            for keyword in keywords:
                current_column_index += 1
                cell = sheet.cell(row=1, column=current_column_index)
                cell.value = keyword
                print(f"Cell Located in {cell.coordinate} Value Updated To -> {cell.value}")
        print("--------------------------------")
        
        default_sheet : str = "Sheet"
        if default_sheet in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[default_sheet])
            
    def append_data(self, sheet_name: str, column_name: str, related: str):
        if related == column_name:
            return
        sheet = self.workbook[sheet_name]
        
        column_index = None
        for index, col in enumerate(sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=1), start=1):
            if sheet.cell(row=1, column=index).value and sheet.cell(row=1, column=index).value.strip() == column_name.strip():
                column_index = index
                break

        if column_index is None:
            print(f"Column '{column_name}' not found in sheet '{sheet_name}'.")
            return

        next_row = 2
        while sheet.cell(row=next_row, column=column_index).value:
            next_row += 1

        sheet.cell(row=next_row, column=column_index).value = related
        print(f"New Value Added: {related} To The Cell -> {sheet.cell(row=next_row, column=column_index).coordinate}")

    
